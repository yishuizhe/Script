import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

# 读取原始Excel文件
input_file = '4.xlsx'  # 替换成你的原始Excel文件路径
df = pd.read_excel(input_file)

# 添加三列数据到DataFrame
df['一审审核是否黑链（总编室）'] = ""
df['二审审核是否黑链（网络安全中心）'] = ""
df['二审备注原因'] = ""

# 获取列数
num_columns = df.shape[1]

# 保存修改后的数据到新的Excel文件
output_file = 'output.xlsx'  # 替换成你的输出Excel文件路径

# 创建一个ExcelWriter对象，用于设置列宽
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Sheet1')  # 指定工作表名称为 'Sheet1'

    # 获取Sheet对象
    wb = writer.book
    ws = wb['Sheet1']

    # 设置列宽
    column_widths = {'黑链': 38, '黑词': 78, '一审审核是否黑链（总编室）': 33, '二审审核是否黑链（网络安全中心）': 41, '二审备注原因': 44}
    for i, (column_name, width) in enumerate(column_widths.items()):
        column_letter = chr(ord('A') + i)  # 获取列字母
        ws.column_dimensions[column_letter].width = width

# 打开新的Excel文件以设置边框
wb = load_workbook(output_file)
ws = wb.active

# 设置边框样式
border_style = Border(
    left=Side(border_style="thin"),
    right=Side(border_style="thin"),
    top=Side(border_style="thin"),
    bottom=Side(border_style="thin")
)

# 应用边框样式到所有单元格
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.border = border_style

# 保存新的Excel文件
wb.save(output_file)

print(f"已保存结果到 {output_file}")
